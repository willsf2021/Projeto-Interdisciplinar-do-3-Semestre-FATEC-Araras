import json
from openpyxl import load_workbook

TACO = load_workbook('taco.xlsx')

def getTitles(sheet, min_col, max_col):
    """Extrai os títulos das colunas de uma planilha"""
    row_title = sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=2, max_row=2, values_only=True)
    titlesTupla = [row for row in row_title]
    titles = titlesTupla[0]
    return titles

def getGrupo(sheet, titles, min_row, max_row, min_col, max_col):
    """Extrai dados de um grupo de alimentos"""
    lista = []
    
    rowData = sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=True)
    dataTuples = [row for row in rowData]

    for data in dataTuples:
        dicio = dict(zip(titles, data))
        lista.append(dicio)
    return lista

def gerarTodosOsDados(grupos_list):
    """Junta todos os grupos em uma única lista"""
    todos_dados = []
    for grupo in grupos_list:
        todos_dados.extend(grupo)
    return todos_dados

# ===== GRUPOS DE ALIMENTOS (ID + NOME) =====
print("Processando grupos de alimentos...")
sheet_grupos = TACO['vitaminas']
titles_grupos = getTitles(sheet_grupos, 1, 2)

grupos_ranges = {
    "Cereais e Derivados": (5, 67),
    "Verduras, Hortaliças e Derivados": (73, 171),
    "Frutas e Derivados": (174, 269),
    "Gorduras e Óleos": (272, 285),
    "Pescados e Frutos do Mar": (289, 338),
    "Carnes e Derivados": (341, 463),
    "Leite e Derivados": (466, 489),
    "Bebidas Alcoólicas e Não Alcoólicas": (492, 505),
    "Ovos e Derivados": (508, 514),
    "Produtos Açucarados": (516, 535),
    "Miscelâneas": (538, 546),
    "Outros Alimentos Industrializados": (549, 553),
    "Alimentos Preparados": (557, 588),
    "Leguminosas e Derivados": (591, 620),
    "Nozes e Sementes": (623, 633)
}

grupos_mongo = []
for nome_grupo, (min_row, max_row) in grupos_ranges.items():
    lista_dados = getGrupo(sheet_grupos, titles_grupos, min_row, max_row, 1, 2)
    grupos_mongo.append({
        "grupo": nome_grupo,
        "alimentos": lista_dados
    })

with open("grupos_alimentos.json", "w", encoding="utf-8") as f:
    json.dump(grupos_mongo, f, ensure_ascii=False, indent=4)

# ===== VITAMINAS =====
print("Processando vitaminas...")
sheet_vitaminas = TACO['vitaminas']
titles_vitaminas = getTitles(sheet_vitaminas, 3, 29)

grupos_vitaminas = []
for nome_grupo, (min_row, max_row) in grupos_ranges.items():
    grupo_dados = getGrupo(sheet_vitaminas, titles_vitaminas, min_row, max_row, 3, 29)
    grupos_vitaminas.extend(grupo_dados)

with open("alimentos_geral.json", "w", encoding="utf-8") as f:
    json.dump(grupos_vitaminas, f, ensure_ascii=False, indent=4)

# ===== AMINOÁCIDOS =====
print("Processando aminoácidos...")
sheet_aminoacidos = TACO['aminoacidos']
titles_aminoacidos = getTitles(sheet_aminoacidos, 3, 21)

aminoacidos_dados = getGrupo(sheet_aminoacidos, titles_aminoacidos, 4, 29, 3, 21)

with open("aminoacidos.json", "w", encoding="utf-8") as f:
    json.dump(aminoacidos_dados, f, ensure_ascii=False, indent=4)

# ===== ÁCIDOS GRAXOS =====
print("Processando ácidos graxos...")
sheet_acidos = TACO['acidos_graxos']
titles_acidos = getTitles(sheet_acidos, 3, 25)

acidos_ranges = {
    "cereais_e_derivados": (5, 61),
    "verduras_hortalicas_e_derivados": (63, 103),
    "frutas_e_derivados": (106, 130),
    "gorduras_e_oleos": (136, 148),
    "pescados_e_frutos_do_mar": (151, 200),
    "carnes_e_derivados": (203, 323),
    "leite_e_derivados": (329, 349),
    "ovos_e_derivados": (352, 357),
    "produtos_acucarados": (363, 372),
    "miscelaneas": (375, 377),
    "outros_alim_indust": (380, 384),
    "alim_preparados": (387, 418),
    "leguminosas_e_derivados": (421, 449),
    "nozes_e_sementes": (452, 461)
}

grupos_acidos = []
for nome_grupo, (min_row, max_row) in acidos_ranges.items():
    grupo_dados = getGrupo(sheet_acidos, titles_acidos, min_row, max_row, 3, 25)
    grupos_acidos.extend(grupo_dados)

with open("acidos_graxos.json", "w", encoding="utf-8") as f:
    json.dump(grupos_acidos, f, ensure_ascii=False, indent=4)
