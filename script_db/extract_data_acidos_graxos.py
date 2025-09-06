import json
from openpyxl import load_workbook

TACO = load_workbook('taco.xlsx')
sheet_vitaminas = TACO['acidos_graxos']


def getTitles():
    row_title = sheet_vitaminas.iter_rows(min_col=3, max_col=25, min_row=2, max_row=2,values_only=True)
    titlesTupla = [row for row in row_title]
    titles = titlesTupla[0]
    return titles

titles = getTitles()

def getGrupo(min_row, max_row):
    lista = []
    
    rowData = sheet_vitaminas.iter_rows(min_col=3, max_col=25, min_row=min_row, max_row=max_row,values_only=True)
    dataTuples = [row for row in rowData]

    for data in dataTuples:
        dicio = dict(zip(titles, data))
        lista.append(dicio)
    return lista

cereais_e_derivados = getGrupo(5, 61)
verduras_hortalicas_e_derivados = getGrupo(63, 103)
frutas_e_derivados = getGrupo(106, 130)
gorduras_e_oleos = getGrupo(136, 148)
pescados_e_frutos_do_mar = getGrupo(151, 200)
carnes_e_derivados = getGrupo(203, 323)
leite_e_derivados = getGrupo(329, 349)
ovos_e_derivados = getGrupo(352, 357)
produtos_acucarados = getGrupo(363, 372)
miscelaneas = getGrupo(375, 377)
outros_alim_indust = getGrupo(380, 384)
alim_preparados = getGrupo(387, 418)
leguminosas_e_derivados = getGrupo(421, 449)
nozes_e_sementes = getGrupo(452, 461)

grupos = [
    cereais_e_derivados,
    verduras_hortalicas_e_derivados,
    frutas_e_derivados,
    gorduras_e_oleos,
    pescados_e_frutos_do_mar,
    carnes_e_derivados,
    leite_e_derivados,
    ovos_e_derivados,
    produtos_acucarados,
    miscelaneas,
    outros_alim_indust,
    alim_preparados,
    leguminosas_e_derivados,
    nozes_e_sementes
]

todos_dados = []
for grupo in grupos:
    todos_dados.extend(grupo)

with open("acidos_graxos.json", "w", encoding="utf-8") as f:
    json.dump(todos_dados, f, ensure_ascii=False, indent=4)