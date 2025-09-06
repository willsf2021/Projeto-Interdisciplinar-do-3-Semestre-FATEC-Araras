import json
from openpyxl import load_workbook

TACO = load_workbook('taco.xlsx')
sheet_vitaminas = TACO['vitaminas']

def getTitles():
    row_title = sheet_vitaminas.iter_rows(min_col=3, max_col=29, min_row=2, max_row=2,values_only=True)
    titlesTupla = [row for row in row_title]
    titles = titlesTupla[0]
    return titles

titles = getTitles()

def getGrupo(min_row, max_row):
    lista = []
    
    rowData = sheet_vitaminas.iter_rows(min_col=3, max_col=29, min_row=min_row, max_row=max_row,values_only=True)
    dataTuples = [row for row in rowData]

    for data in dataTuples:
        dicio = dict(zip(titles, data))
        lista.append(dicio)
    return lista

cereais_e_derivados = getGrupo(5, 67)
verduras_hortalicas_e_derivados = getGrupo(73, 171)
frutas_e_derivados = getGrupo(174, 269)
gorduras_e_oleos = getGrupo(272, 285)
pescados_e_frutos_do_mar = getGrupo(289, 338)
carnes_e_derivados = getGrupo(341, 463)
leite_e_derivados = getGrupo(466, 489)
bebidas_alcoolicas_e_n_alcoolicas = getGrupo(492, 505)
ovos_e_derivados = getGrupo(508, 514)
produtos_acucarados = getGrupo(516, 535)
miscelaneas = getGrupo(538, 546)
outros_alim_indust = getGrupo(549, 553)
alim_preparados = getGrupo(557, 588)
leguminosas_e_derivados = getGrupo(591, 620)
nozes_e_sementes = getGrupo(623, 633)

grupos = [
    cereais_e_derivados,
    verduras_hortalicas_e_derivados,
    frutas_e_derivados,
    gorduras_e_oleos,
    pescados_e_frutos_do_mar,
    carnes_e_derivados,
    leite_e_derivados,
    bebidas_alcoolicas_e_n_alcoolicas,
    ovos_e_derivados,
    produtos_acucarados,
    miscelaneas,
    outros_alim_indust,
    alim_preparados,
    leguminosas_e_derivados,
    nozes_e_sementes
]

# Junta todos os grupos em uma única lista
todos_dados = []
for grupo in grupos:
    todos_dados.extend(grupo)

with open("alimentos_geral.json", "w", encoding="utf-8") as f:
    json.dump(todos_dados, f, ensure_ascii=False, indent=4)