import json
from openpyxl import load_workbook

TACO = load_workbook('taco.xlsx')
sheet_vitaminas = TACO['aminoacidos']


def getTitles():
    row_title = sheet_vitaminas.iter_rows(min_col=3, max_col=21, min_row=2, max_row=2,values_only=True)
    titlesTupla = [row for row in row_title]
    titles = titlesTupla[0]
    return titles

titles = getTitles()

def getGrupo(min_row, max_row):
    lista = []
    
    rowData = sheet_vitaminas.iter_rows(min_col=3, max_col=21, min_row=min_row, max_row=max_row,values_only=True)
    dataTuples = [row for row in rowData]

    for data in dataTuples:
        dicio = dict(zip(titles, data))
        lista.append(dicio)
    return lista

geral = getGrupo(4, 29)


with open("aminoacidos.json", "w", encoding="utf-8") as f:
    json.dump(geral, f, ensure_ascii=False, indent=4)