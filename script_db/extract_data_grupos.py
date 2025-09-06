import json
from openpyxl import load_workbook

TACO = load_workbook('taco.xlsx')
sheet_vitaminas = TACO['vitaminas']


def getTitles():
    row_title = sheet_vitaminas.iter_rows(min_col=1, max_col=2, min_row=2, max_row=2,values_only=True)
    titlesTupla = [row for row in row_title]
    titles = titlesTupla[0]
    return titles

titles = getTitles()

def getGrupo(min_row, max_row):
    lista = []
    
    rowData = sheet_vitaminas.iter_rows(min_col=1, max_col=2, min_row=min_row, max_row=max_row,values_only=True)
    dataTuples = [row for row in rowData]

    for data in dataTuples:
        dicio = dict(zip(titles, data))
        lista.append(dicio)
    return lista

cereais_e_derivados = getGrupo(5, 67)
verduras_hortalicas_e_derivados = getGrupo(73, 171)
frutas_e_derivados = getGrupo(174, 269)
gorduras_e_oleos = getGrupo(272, 285)
pescados_e_frutos_do_mar = getGrupo(289, 338)
carnes_e_derivados = getGrupo(341, 463)
leite_e_derivados = getGrupo(466, 489)
bebidas_alcoolicas_e_n_alcoolicas = getGrupo(492, 505)
ovos_e_derivados = getGrupo(508, 514)
produtos_acucarados = getGrupo(516, 535)
miscelaneas = getGrupo(538, 546)
outros_alim_indust = getGrupo(549, 553)
alim_preparados = getGrupo(557, 588)
leguminosas_e_derivados = getGrupo(591, 620)
nozes_e_sementes = getGrupo(623, 633)

grupos = {
    "Cereais e Derivados": cereais_e_derivados,
    "Verduras, Hortaliças e Derivados": verduras_hortalicas_e_derivados,
    "Frutas e Derivados": frutas_e_derivados,
    "Gorduras e Óleos": gorduras_e_oleos,
    "Pescados e Frutos do Mar": pescados_e_frutos_do_mar,
    "Carnes e Derivados": carnes_e_derivados,
    "Leite e Derivados": leite_e_derivados,
    "Bebidas Alcoólicas e Não Alcoólicas": bebidas_alcoolicas_e_n_alcoolicas,
    "Ovos e Derivados": ovos_e_derivados,
    "Produtos Açucarados": produtos_acucarados,
    "Miscelâneas": miscelaneas,
    "Outros Alimentos Industrializados": outros_alim_indust,
    "Alimentos Preparados": alim_preparados,
    "Leguminosas e Derivados": leguminosas_e_derivados,
    "Nozes e Sementes": nozes_e_sementes
}

grupos_mongo = []

for nome_grupo, lista_dados in grupos.items():
    grupos_mongo.append({
        "grupo": nome_grupo,
        "alimentos": lista_dados 
    })

# Salva em JSON
with open("grupos_alimentos.json", "w", encoding="utf-8") as f:
    json.dump(grupos_mongo, f, ensure_ascii=False, indent=4)

print("Arquivo 'grupos_alimentos.json' criado com sucesso!")